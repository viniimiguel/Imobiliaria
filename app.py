from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
import openpyxl
from selenium.common.exceptions import NoSuchElementException
import re




class Imovel():
    def __init__(self):
        self.driver = webdriver.Chrome()
        self.site_link = 'https://www.zapimoveis.com.br/venda/imoveis/pe+recife/?__ab=exp-aa-test:B,new-discover-zap:alert,vas-officialstore-social:enabled,deduplication:control&transacao=venda&onde=,Pernambuco,Recife,,,,,city,BR>Pernambuco>NULL>Recife,-8.05774,-34.882963,&pagina=1'
        
    def main(self):
        self.entra()
        sleep(5)
        self.raspagem()
        sleep(5)
        self.cria_planilhas()
        sleep(12831830)


    def entra(self):
        self.driver.get(self.site_link)
        self.driver.maximize_window()

    def raspagem(self):
        global armazena_logradouro, armazena_rua, armazena_descricao, armazena_metragem, armazena_comodos, armazena_banheiros, armazena_vagas, armazena_valor

        armazena_logradouro = []
        armazena_rua = []
        armazena_descricao = []
        armazena_metragem = []
        armazena_comodos = []
        armazena_banheiros = []
        armazena_vagas = []
        armazena_valor = []
        
        contador = 1

        
        while True:
            try:
                xpats = {
                    'XP': {
                        'logradouro': f'/html/body/div[1]/main/section/div/form/div[2]/div[2]/div[1]/div[2]/div[{contador}]/div/a/div/div/div[2]/div[1]/section/div/h2',
                        'rua': f'/html/body/div[1]/main/section/div/form/div[2]/div[2]/div[1]/div[2]/div[{contador}]/div/a/div/div/div[2]/div[1]/section/p',
                        'descricao': f'/html/body/div[1]/main/section/div/form/div[2]/div[2]/div[1]/div[2]/div[{contador}]/div/a/div/div/div[2]/div[2]/p',
                        'info': f'/html/body/div[1]/main/section/div/form/div[2]/div[2]/div[1]/div[2]/div[{contador}]/div/a/div/div/div[2]/section',
                        'valor': f'/html/body/div[1]/main/section/div/form/div[2]/div[2]/div[1]/div[2]/div[{contador}]/div/a/div/div/div[2]/div[3]/div[1]/p',
                        'button_next': f'/html/body/div[1]/main/section/div/form/div[2]/div[2]/div[1]/div[2]/section/nav/button[2]',
                    }
                }

                logradouro = self.driver.find_element(By.XPATH, xpats['XP']['logradouro'])
                armazena_logradouro.append(logradouro.text)

                rua = self.driver.find_element(By.XPATH, xpats['XP']['rua'])
                armazena_rua.append(rua.text)

                descricao = self.driver.find_element(By.XPATH, xpats['XP']['descricao'])
                armazena_descricao.append(descricao.text)

                info = self.driver.find_element(By.XPATH, xpats['XP']['info'])

                partes = info.text.split()

                if len(partes) >= 4:
                    metragems = partes[0:2]
                    metragem = " ".join(metragems)
                    comodos = partes[2]
                    banheiros = partes[3]
                    vagas = partes[4] if len(partes) > 4 else '0'

                    print("Metragem:", metragem)
                    print("Comodos:", comodos)
                    print("Banheiros:", banheiros)
                    print("Vagas:", vagas)
                    armazena_metragem.append(metragem)
                    armazena_comodos.append(comodos)
                    armazena_banheiros.append(banheiros)
                    armazena_vagas.append(vagas)
                else:
                    print("A string não contém informações suficientes.")
                

                valor = self.driver.find_element(By.XPATH, xpats['XP']['valor'])
                armazena_valor.append(valor.text)

                self.driver.execute_script('arguments[0].scrollIntoView();', logradouro)

                print(logradouro.text)
                print(rua.text)
                print(descricao.text)
                print(info.text)
                print(valor.text)
                print('###'*40)

                contador += 1

            except NoSuchElementException:
                try:
                    button_next = self.driver.find_element(By.XPATH, xpats['XP']['button_next'])
                    if button_next:
                        button_next.click()
                        contador = 1
                        sleep(5)
                except NoSuchElementException:
                    print('raspagem finalizada')
                    break
                except Exception as e:
                    print('ops!!! ocorreu um error inesperado {e}')
                    break
    
    def cria_planilhas(self):
        planilha = openpyxl.Workbook()
        imovel = planilha.active
        imovel.title = 'Imoveis'
        imovel['A1'] = 'LOGRADOURO'
        imovel['B1'] = 'RUA'
        imovel['C1'] = 'DESCRIÇÃO'
        imovel['D1'] = 'METRAGEM'
        imovel['E1'] = 'COMODOS'
        imovel['F1'] = 'BANHEIROS'
        imovel['G1'] = 'VAGAS'
        imovel['H1'] = 'VALOR'

        for index, (logradouro, rua, descricao, metragem, comodos, banheiros, vagas, valor) in enumerate(zip(armazena_logradouro, armazena_rua, armazena_descricao, armazena_metragem, armazena_comodos, armazena_banheiros, armazena_vagas, armazena_valor), start=2):
            imovel.cell(column=1, row=index, value=logradouro)
            imovel.cell(column=2, row=index, value=rua)
            imovel.cell(column=3, row=index, value=descricao)
            imovel.cell(column=4, row=index, value=metragem)
            imovel.cell(column=5, row=index, value=comodos)
            imovel.cell(column=6, row=index, value=banheiros)
            imovel.cell(column=7, row=index, value=vagas)
            imovel.cell(column=8, row=index, value=valor)

        planilha.save('planilha_de_precos_atacadao.xlsx')
        print('planilha salva com sucesso!')


imovel = Imovel()
imovel.main()